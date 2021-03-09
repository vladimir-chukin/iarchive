#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Название: iarchive.py
# Описание: Программа для работы с архивом данных archive.org, управляемая ключами командной строки
# ТЗ:       https://docs.google.com/document/d/1KMWpKWA96OurH2V9oIm2444IJM2iO5BPi_QwrjAf0FQ/edit
#
# Установка библиотек:
#  pip install internetarchive
#  pip install pandas
#  pip install openpyxl

from argparse import ArgumentParser
from openpyxl import load_workbook
from sys import getdefaultencoding
from email.utils import parseaddr
import pandas as pd
import subprocess
import json
import sys
import os
import re


#-----------------------------------------------------------------------
# чтение списка запросов из XLSX-файла
#-----------------------------------------------------------------------
def read_queries(in_file_name):
    try:
        df = pd.read_excel(in_file_name, dtype={'query': str, 'dir_name': str})
        return dict(zip(df['query'].tolist(), df['dir_name'].tolist()))
    except:
        return {}


#-----------------------------------------------------------------------
# поиск идентификаторов по запросу
#-----------------------------------------------------------------------
def search_identifiers(query):
    result = subprocess.run(['ia', 'search', query, '--itemlist'], stdout=subprocess.PIPE)
    list_data = result.stdout.decode('utf-8').splitlines()
    return list_data


#-----------------------------------------------------------------------
# получение метаданных по идентификатору
#-----------------------------------------------------------------------
def get_dict_data(identifier):
    result = subprocess.run(['ia', 'metadata', identifier], stdout=subprocess.PIPE)
    json_data = result.stdout.decode('utf-8')
    dict_data = json.loads(json_data)
    return dict_data


#-----------------------------------------------------------------------
# получение значения свойства из метаданных
#-----------------------------------------------------------------------
def get_property(meta_data, prop):
    if prop in meta_data.keys() and len(meta_data[prop]) > 0:
        return meta_data[prop]
    return ''


#-----------------------------------------------------------------------
# замена недопустимых символов в имени файла (Windows)
#-----------------------------------------------------------------------
def check_file_name(file_name, wildcard=''):

    # задаем строку с недопустимыми символами
    invalid = '<>^:;"/\|?*+@#$.'

    # удаляем недопустимые символы
    for char in invalid:
        file_name = file_name.replace(char, wildcard)

    # устанавливаем кодировку символов по умолчанию
    getdefaultencoding()

    # проверяем символы на соответствие кодировке URF-8
    file_name = file_name.encode().decode('utf-8', errors='surrogateescape')

    # удаляем пробелы в начале и конце названия
    file_name = file_name.strip(' ')

    # заменяем несколько пробелов одним пробелом
    file_name = re.sub(' +', ' ', file_name)

    return file_name


#-----------------------------------------------------------------------
# определение языка в соотвествии со стандартом ISO 639-2 (три символа)
#-----------------------------------------------------------------------
def iso_639_2(language):

    # преобразование списка в строку
    if isinstance(language, list):
        language = language[0]

    # создаем словарь
    lang_dict = {'Russian': 'rus', 'Greek': 'gre', 'Greek, Ancient': 'grc', 'Latin': 'lat',
                 'Slavic languages': 'sla', 'Church Slavic': 'chu', 'Georgian': 'kat',
                 'Bulgarian': 'bul', 'Armenian': 'arm', 'Serbian': 'srp',
                 'English': 'eng', 'German': 'deu', 'Spanish': 'spa', 'Italian': 'ita',
                 'Portuguese': 'por', 'French': 'fre'}

    if language in lang_dict.values():
        return language
    elif language in lang_dict.keys():
        return lang_dict[language]
    else:
        return ''


#-----------------------------------------------------------------------
# парсинг адреса электронной почты из строки
#-----------------------------------------------------------------------
def email(text):
    return parseaddr(text)[1]


#-----------------------------------------------------------------------
# поиск информации по списку запросов и запись в файл
#-----------------------------------------------------------------------
def search_and_save(queries, types, out_file_name):

    # задаем имя листа в XLSX-файле
    sheet_name = 'Sheet1'

    # создаем датафрейм
    columns = ['query', 'identifier', 'title', 'creator', 'year', 'uploader', 'uploader_email',
               'language', 'in_file_path', 'in_file_name', 'out_file_path', 'out_file_name']
    df = pd.DataFrame(columns=columns)

    # сохраняем заголовок в XLSX-файл
    writer = pd.ExcelWriter(out_file_name, engine='openpyxl', mode='w')
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

    # перебираем все запросы
    for q, (query, out_file_path) in enumerate(queries.items()):

        # получаем список идентификаторов для данного запроса
        identifiers = search_identifiers(query)

        # если индентификатор уже встречался при выполнении предыдущих запросов, то исключаем его из списка
        identifiers_processed = pd.read_excel(out_file_name)['identifier'].tolist()
        identifiers = [identifier for identifier in identifiers if identifier not in identifiers_processed]

        # перебираем по очереди все идентификаторы
        for i, identifier in enumerate(identifiers):

            # название произведения
            book_title = ''

            # получаем словарь с данными для указанного идентификатора
            dict_data = get_dict_data(identifier)

            if 'metadata' in dict_data:
                meta_data = dict_data['metadata']
                row = {}
                row['query'] = query
                row['identifier'] = identifier
                row['title'] = get_property(meta_data, 'title')
                row['creator'] = get_property(meta_data, 'creator')
                row['year'] = get_property(meta_data, 'year')
                uploader = get_property(meta_data, 'uploader')
                row['uploader'] = uploader
                row['uploader_email'] = email(uploader)
                # записываем язык в соотвествии со стандартом ISO 639-2 (три символа)
                row['language'] = iso_639_2(get_property(meta_data, 'language'))
                row['out_file_path'] = out_file_path

                # определяем авторов
                authors = ''
                if isinstance(row['creator'], list):
                    for c, creator in enumerate(row['creator']):
                        if c > 0:
                            authors += ', '
                        authors += creator.split(',')[0]
                else:
                    authors += row['creator'].split(',')[0]
                authors = check_file_name(authors)

                # определяем год издания
                if isinstance(row['year'], list):
                    year = row['year'][-1]
                else:
                    year = row['year']
                year = check_file_name(year)

                # ограничиваем длину названия файла
                title = check_file_name(row['title'])
                max_title_len = 127 - 6 - 7 - len(authors) - len(year)
                if len(title) > max_title_len:
                    title = title[:max_title_len].strip(' ')

                # формируем название файла без расширения
                if len(row['title']) > 0:
                    if len(year) > 0 and len(authors) > 0:
                        book_title = '%s, %s (%s)' % (title, authors, year)
                    elif len(year) > 0 and len(authors) == 0:
                        book_title = '%s, (%s)' % (title, year)
                    elif len(year) == 0 and len(authors) > 0:
                        book_title = '%s, %s' % (title, authors)
                    else:
                        book_title = '%s' % (title)
                    row['out_file_name'] = book_title
                else:
                    continue
            else:
                continue

            if 'dir' in dict_data:
                row['in_file_path'] = dict_data['dir']
            else:
                row['in_file_path'] = ''

            if 'files' in dict_data:
                files_data = dict_data['files']
                for file_data in files_data:
                    try:
                        # имя файла
                        file_name = file_data['name']
                        # имя файла без расширения
                        name = file_name.split('.')[-2]
                        # расширение файла
                        extention = file_name.split('.')[-1]
                        # список суффиксов в именах файлов, которые следует пропускать
                        #suffixes = ['_text', '_encrypted', '_bw']
                        suffixes = []
                        # если имя файла не содержит суффиксов и соответствует одному из типов, то сохраняем
                        if not any([suffix in name for suffix in suffixes]) and (len(types) == 0 or (len(types) > 0 and extention in types)):
                            row['in_file_name'] = file_name
                            count = 0
                            for fn in list(df['out_file_name']):
                                count += fn.count(extention)
                            if count > 0:
                                row['out_file_name'] = book_title + ' (%d)'%count + '.' + extention
                            else:
                                row['out_file_name'] = book_title + '.' + extention

                            # отображаем номер, общее количество и текст запроса, порядковый номер и общее количество
                            # идентификаторов, новое имя файла
                            print('[%d/%d] %s [%d/%d] %s' % (q+1, len(queries), query, i+1, len(identifiers), row['out_file_name']))

                            # добавляем запись в датафрейм
                            df = df.append(row, ignore_index=True)
                    except:
                        pass

            # записываем результат в XLSX-файл
            if len(df) > 0:
                writer = pd.ExcelWriter(out_file_name, engine='openpyxl', mode='a')
                writer.book = load_workbook(filename=out_file_name)
                # создаем список листов в XLSX-файле
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                # определяем число строк в XLSX-файле
                startrow = writer.book[sheet_name].max_row
                # записываем данные в XLSX-файл
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=None)
                writer.save()
                # очищаем датафрейм
                df = df.drop([x for x in range(len(df))], axis=0)

        # загружаем данные из текущего XLSX-файла
        dfs = pd.read_excel(out_file_name)
        # сортируем записи по названиям запросов и именам новых файлов
        dfs = dfs.sort_values(by=['query', 'out_file_name'])
        # записываем отсортированные данные обратно в XLSX-файл
        writer = pd.ExcelWriter(out_file_name, engine='openpyxl', mode='w')
        dfs.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()


#-----------------------------------------------------------------------
# вывод справки на консоль
#-----------------------------------------------------------------------
def print_help(param=None):
    if param == None:
        print('usage: iarchive.py [-h] [-s [SEARCH [SEARCH ...]]] [-i [INPUT]] [-o [OUTPUT]]')
        print('                   [-t [TYPES [TYPES ...]]]')
        print()
        print('optional arguments:')
        print('  -h, --help            show this help message and exit')
        print('  -s [SEARCH [SEARCH ...]], --search [SEARCH [SEARCH ...]]')
        print('                        mode of forming the list of downloadable files')
        print('  -i [INPUT], --input [INPUT]')
        print('                        name of XLSX file with the list of queries')
        print('  -o [OUTPUT], --output [OUTPUT]')
        print('                        name of XLSX file with the list of downloadable files')
        print('  -t [TYPES [TYPES ...]], --types [TYPES [TYPES ...]]')
        print('                        file name extensions')
    else:
        if param=='-s':
            print('example #1:\n iarchive.py -s "Edward Pusey"')
            print('example #2:\n iarchive.py -s "Edward Pusey", "Church Fathers", "Philocalia"')
            print('example #3:\n iarchive.py -s "Edward Pusey", "Church Fathers", Philocalia')
        if param=='--search':
            print('example #1:\n iarchive.py --search "Edward Pusey"')
            print('example #2:\n iarchive.py --search "Edward Pusey", "Church Fathers", "Philocalia"')
            print('example #3:\n iarchive.py --search "Edward Pusey", "Church Fathers", Philocalia')
        if param=='-i':
            print('example #1:\n iarchive.py --search "Edward Pusey" -i')
            print('example #2:\n iarchive.py --search "Edward Pusey" -i query_2.xlsx')
        if param=='--input':
            print('example #1:\n iarchive.py --search "Edward Pusey" --input')
            print('example #2:\n iarchive.py --search "Edward Pusey" --input query_2.xlsx')
        if param=='-o':
            print('example #1:\n iarchive.py --search "Edward Pusey" -o')
            print('example #2:\n iarchive.py --search "Edward Pusey" -o download_2.xlsx')
        if param=='--output':
            print('example #1:\n iarchive.py --search "Edward Pusey" --output')
            print('example #2:\n iarchive.py --search "Edward Pusey" --output download_2.xlsx')
        if param=='-t':
            print('example #1:\n iarchive.py --search -t pdf djvu mp3 ogg')
            print('example #2:\n iarchive.py --search -t pdf, djvu, mp3, ogg')
            print('example #3:\n iarchive.py --search -t "pdf djvu mp3 ogg"')
            print('example #4:\n iarchive.py --search -t "pdf", "djvu", "mp3", "ogg"')
        if param=='--types':
            print('example #1:\n iarchive.py --search --types pdf djvu mp3 ogg')
            print('example #2:\n iarchive.py --search --types pdf, djvu, mp3, ogg')
            print('example #3:\n iarchive.py --search --types "pdf djvu mp3 ogg"')
            print('example #4:\n iarchive.py --search --types "pdf", "djvu", "mp3", "ogg"')


if __name__ == '__main__':

    # задаем имя файла со списком запросов по умолчанию
    in_file_name = 'query.xlsx'

    # задаем имя файла со списком загружаемых файлов по умолчанию
    out_file_name = 'download.xlsx'

    # вывод справочной информации
    if len(sys.argv)==2 and ('-h' in sys.argv[1] or '--help' in sys.argv[1] or '/?' in sys.argv[1]):
        print_help()
        exit(0)
    if len(sys.argv)>=3 and ('-h' in sys.argv[1] or '--help' in sys.argv[1] or '/?' in sys.argv[1]):
        print_help(sys.argv[2])
        exit(0)

    # разбор параметров командной строки
    parser = ArgumentParser()
    parser.add_argument('-s', '--search', nargs='*', help='mode of forming the list of downloadable files')
    parser.add_argument('-i', '--input', nargs='?', const=in_file_name, help='name of XLSX file with the list of queries')
    parser.add_argument('-o', '--output', nargs='?', const=out_file_name, help='name of XLSX file with the list of downloadable files')
    parser.add_argument('-t', '--types', nargs='*', help='file name extensions')
    args = parser.parse_args()

    # режим поиска файлов по запросам
    if args.search is not None:

        # список запросов
        queries = {}

        # загружаем запросы из командной строки
        if len(args.search) > 0:
            # создаем список
            queries_list = [query[:-1] if query[-1] == ',' else query for query in args.search]
            # создаем словарь
            queries_dict = {query: check_file_name(query, ' ') for query in queries_list}
            # добавляем в словарь
            queries = dict(list(queries_dict.items()) + list(queries.items()))


        # если указан ключ --input и файл существует, то считываем из него список запросов
        if args.input and os.path.isfile(args.input):
            queries = dict(list(read_queries(args.input).items()) + list(queries.items()))

        # если заданы расширения файлов, то сохраняем их в списке
        #types = ['pdf', 'djvu', 'mp3', 'ogg']
        types = []
        if args.types:
            types = []
            if len(args.types) == 1:
                types.append( [tp.split(' ') if len(tp.split(' ')) > 1 else tp for tp in args.types][0] )
            else:
                types += [tp[:-1] if tp[-1] == ',' else tp for tp in args.types]

        # если указан ключ --output, то запускаем процесс выполнения запросов и запись результатов в выходной файл
        if args.output:
            search_and_save(queries, types, args.output)
    else:
        print_help()
